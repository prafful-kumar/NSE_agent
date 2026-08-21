from __future__ import annotations

"""Unit tests for pnl_report_parser.py. Synthetic workbooks mirror the real
Zerodha Console P&L/holdings export structure (confirmed by inspecting real
files under statements/ -- never used directly as test fixtures, same
discipline as test_broker_history_zerodha_importer.py)."""

from decimal import Decimal

import openpyxl
import pytest

from investing_agent.services.reconstruction.pnl_report_parser import (
    HOLDINGS_EQUITY_HEADER,
    PNL_EQUITY_HEADER,
    parse_holdings_statement,
    parse_pnl_statement,
)


def _write_pnl_workbook(path, rows, period="2025-08-20 to 2026-08-19"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equity"
    ws.append((None, "Client ID", "TEST"))
    ws.append((None, f"P&L Statement for Equity from {period}"))
    ws.append((None, "Summary"))
    ws.append((None, "Realized P&L", -100.5))
    ws.append((None, "Unrealized P&L", 200.0))
    ws.append((None, *PNL_EQUITY_HEADER))
    for row in rows:
        ws.append((None, *row))
    wb.save(path)


def _write_holdings_workbook(path, rows, as_of="2026-08-19"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equity"
    ws.append((None, "Client ID", "TEST"))
    ws.append((None, f"Equity Holdings Statement as on {as_of}"))
    ws.append((None, "Summary"))
    ws.append((None, "Invested Value", 1000.0))
    ws.append((None, "Present Value", 1200.0))
    ws.append((None, "Unrealized P&L", 200.0))
    ws.append((None, *HOLDINGS_EQUITY_HEADER))
    for row in rows:
        ws.append((None, *row))
    wb.save(path)


class TestPnlStatementParsing:
    def test_parses_period_and_rows(self, tmp_path) -> None:
        path = tmp_path / "pnl-TEST.xlsx"
        _write_pnl_workbook(
            path,
            [("INFY", "INE009A01021", 0.0, 0.0, 0.0, 0.0, 0.0, 1500.0, 10.0, "", 15000.0, 500.0, 3.33)],
        )
        stmt = parse_pnl_statement(path)

        assert stmt.period_start.isoformat() == "2025-08-20"
        assert stmt.period_end.isoformat() == "2026-08-19"
        assert stmt.summary_realized_pnl == Decimal("-100.5")
        assert stmt.summary_unrealized_pnl == Decimal("200.0")
        assert len(stmt.rows) == 1
        row = stmt.rows[0]
        assert row.symbol == "INFY"
        assert row.open_quantity == Decimal("10.0")
        assert row.unrealized_pnl == Decimal("500.0")

    def test_raises_on_unrecognized_header(self, tmp_path) -> None:
        path = tmp_path / "pnl-TEST.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append((None, "Symbol", "Something", "Else"))
        wb.save(path)

        with pytest.raises(ValueError, match="Unrecognized P&L statement format"):
            parse_pnl_statement(path)


class TestHoldingsStatementParsing:
    def test_parses_as_of_date_and_rows(self, tmp_path) -> None:
        path = tmp_path / "holdings-TEST.xlsx"
        _write_holdings_workbook(
            path,
            [
                ("INFY", "INE009A01021", "IT", 10.0, 0.0, 10.0, 0.0, 0.0, 1400.0, 1500.0, 1000.0, 7.14),
            ],
        )
        stmt = parse_holdings_statement(path)

        assert stmt.as_of_date.isoformat() == "2026-08-19"
        assert stmt.invested_value == Decimal("1000.0")
        assert len(stmt.rows) == 1
        row = stmt.rows[0]
        assert row.symbol == "INFY"
        assert row.quantity_available == Decimal("10.0")
        assert row.average_price == Decimal("1400.0")

    def test_raises_on_unrecognized_header(self, tmp_path) -> None:
        path = tmp_path / "holdings-TEST.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append((None, "Symbol", "Something", "Else"))
        wb.save(path)

        with pytest.raises(ValueError, match="Unrecognized holdings statement format"):
            parse_holdings_statement(path)
