from __future__ import annotations

"""Parses Zerodha Console P&L and holdings exports (Phase 6B).

These are NOT import sources (see zerodha.py's module docstring) -- they
are read-only cross-check inputs for reconciliation.py, never written to
historical_trades/cash_flows. Header rows are located dynamically by exact
column-signature match, same defensive pattern as zerodha.py, since these
exports share the same Console template (metadata rows of varying length,
then a header row, then data).
"""

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

PNL_EQUITY_HEADER: tuple[str, ...] = (
    "Symbol",
    "ISIN",
    "Quantity",
    "Buy Value",
    "Sell Value",
    "Realized P&L",
    "Realized P&L Pct.",
    "Previous Closing Price",
    "Open Quantity",
    "Open Quantity Type",
    "Open Value",
    "Unrealized P&L",
    "Unrealized P&L Pct.",
)

HOLDINGS_EQUITY_HEADER: tuple[str, ...] = (
    "Symbol",
    "ISIN",
    "Sector",
    "Quantity Available",
    "Quantity Discrepant",
    "Quantity Long Term",
    "Quantity Pledged (Margin)",
    "Quantity Pledged (Loan)",
    "Average Price",
    "Previous Closing Price",
    "Unrealized P&L",
    "Unrealized P&L Pct.",
)


def _trimmed_row(row: tuple) -> tuple:
    values = list(row[1:])
    while values and values[-1] is None:
        values.pop()
    return tuple(values)


def _find_header_row(ws, signature: tuple[str, ...]) -> int | None:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _trimmed_row(row) == signature:
            return row_idx
    return None


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"Cannot parse decimal value: {value!r}") from exc


def _find_title_row(ws, prefix: str) -> str | None:
    for row in ws.iter_rows(values_only=True):
        values = _trimmed_row(row)
        if values and isinstance(values[0], str) and values[0].startswith(prefix):
            return values[0]
    return None


def _find_summary_value(ws, label: str) -> Decimal | None:
    for row in ws.iter_rows(values_only=True):
        values = _trimmed_row(row)
        if len(values) >= 2 and values[0] == label:
            return _to_decimal(values[1])
    return None


@dataclass
class PnlStatementRow:
    symbol: str
    isin: str | None
    quantity: Decimal
    buy_value: Decimal
    sell_value: Decimal
    realized_pnl: Decimal
    previous_closing_price: Decimal
    open_quantity: Decimal
    open_value: Decimal
    unrealized_pnl: Decimal


@dataclass
class PnlStatement:
    file_name: str
    period_start: date
    period_end: date
    rows: list[PnlStatementRow]
    summary_realized_pnl: Decimal
    summary_unrealized_pnl: Decimal


@dataclass
class HoldingsRow:
    symbol: str
    isin: str | None
    quantity_available: Decimal
    average_price: Decimal
    previous_closing_price: Decimal
    unrealized_pnl: Decimal


@dataclass
class HoldingsStatement:
    file_name: str
    as_of_date: date
    rows: list[HoldingsRow]
    invested_value: Decimal
    present_value: Decimal
    unrealized_pnl: Decimal


def parse_pnl_statement(file_path: Path) -> PnlStatement:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Equity"] if "Equity" in wb.sheetnames else wb[wb.sheetnames[0]]

    title = _find_title_row(ws, "P&L Statement for Equity from ")
    if title is None:
        raise ValueError(
            f"Unrecognized P&L statement format in {file_path.name!r}: no "
            "'P&L Statement for Equity from ...' title row found."
        )
    # "P&L Statement for Equity from 2025-08-20 to 2026-08-19"
    date_part = title.removeprefix("P&L Statement for Equity from ")
    start_str, end_str = date_part.split(" to ")
    period_start = date.fromisoformat(start_str.strip())
    period_end = date.fromisoformat(end_str.strip())

    header_row = _find_header_row(ws, PNL_EQUITY_HEADER)
    if header_row is None:
        raise ValueError(
            f"Unrecognized P&L statement format in {file_path.name!r}: header row "
            f"did not match expected columns {PNL_EQUITY_HEADER}."
        )

    rows: list[PnlStatementRow] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = _trimmed_row(row)
        if not values:
            continue
        (
            symbol, isin, quantity, buy_value, sell_value, realized_pnl,
            _realized_pnl_pct, previous_closing_price, open_quantity,
            _open_quantity_type, open_value, unrealized_pnl, _unrealized_pnl_pct,
        ) = values
        rows.append(
            PnlStatementRow(
                symbol=str(symbol),
                isin=str(isin) if isin else None,
                quantity=_to_decimal(quantity),
                buy_value=_to_decimal(buy_value),
                sell_value=_to_decimal(sell_value),
                realized_pnl=_to_decimal(realized_pnl),
                previous_closing_price=_to_decimal(previous_closing_price),
                open_quantity=_to_decimal(open_quantity),
                open_value=_to_decimal(open_value),
                unrealized_pnl=_to_decimal(unrealized_pnl),
            )
        )

    return PnlStatement(
        file_name=file_path.name,
        period_start=period_start,
        period_end=period_end,
        rows=rows,
        summary_realized_pnl=_find_summary_value(ws, "Realized P&L") or Decimal("0"),
        summary_unrealized_pnl=_find_summary_value(ws, "Unrealized P&L") or Decimal("0"),
    )


def parse_holdings_statement(file_path: Path) -> HoldingsStatement:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Equity"] if "Equity" in wb.sheetnames else wb[wb.sheetnames[0]]

    title = _find_title_row(ws, "Equity Holdings Statement as on ")
    if title is None:
        raise ValueError(
            f"Unrecognized holdings statement format in {file_path.name!r}: no "
            "'Equity Holdings Statement as on ...' title row found."
        )
    as_of_date = date.fromisoformat(title.removeprefix("Equity Holdings Statement as on ").strip())

    header_row = _find_header_row(ws, HOLDINGS_EQUITY_HEADER)
    if header_row is None:
        raise ValueError(
            f"Unrecognized holdings statement format in {file_path.name!r}: header "
            f"row did not match expected columns {HOLDINGS_EQUITY_HEADER}."
        )

    rows: list[HoldingsRow] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = _trimmed_row(row)
        if not values:
            continue
        (
            symbol, isin, _sector, quantity_available, _quantity_discrepant,
            _quantity_long_term, _quantity_pledged_margin, _quantity_pledged_loan,
            average_price, previous_closing_price, unrealized_pnl, _unrealized_pnl_pct,
        ) = values
        rows.append(
            HoldingsRow(
                symbol=str(symbol),
                isin=str(isin) if isin else None,
                quantity_available=_to_decimal(quantity_available),
                average_price=_to_decimal(average_price),
                previous_closing_price=_to_decimal(previous_closing_price),
                unrealized_pnl=_to_decimal(unrealized_pnl),
            )
        )

    return HoldingsStatement(
        file_name=file_path.name,
        as_of_date=as_of_date,
        rows=rows,
        invested_value=_find_summary_value(ws, "Invested Value") or Decimal("0"),
        present_value=_find_summary_value(ws, "Present Value") or Decimal("0"),
        unrealized_pnl=_find_summary_value(ws, "Unrealized P&L") or Decimal("0"),
    )
