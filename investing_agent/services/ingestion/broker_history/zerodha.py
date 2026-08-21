from __future__ import annotations

"""Zerodha Console statement importer (Phase 6A).

Built against real Console exports inspected in `statements/`:
tradebook-NP7356-EQ.xlsx and ledger-NP7356.xlsx. Both files share the same
Console export template: 14 metadata/title rows, then a header row, then
data rows. The header row is located dynamically by matching a known exact
column signature -- never assumed to be at a fixed row index -- so a
slightly different metadata block (longer title string, etc.) in a future
export does not silently misparse.

Console's P&L report and holdings statement were also inspected. Neither is
parsed here:
  - P&L report ("pnl-*.xlsx") gives a per-symbol AGGREGATE realized P&L over
    the export's date range, not per-trade facts -- it has no place in the
    HistoricalTrade/CashFlow/Dividend fact tables. It's a natural
    cross-check for Phase 6B's FIFO lot reconstruction, not a Phase 6A
    import source.
  - Holdings statement ("holdings-*.xlsx") is a point-in-time snapshot ("as
    on <today>"), not a historical time series. Same future use: validating
    Phase 6B's reconstructed-as-of-today holdings, not something to import
    now.

Dividends: no dividend entries exist in the inspected ledger export, and
the ledger has no symbol/quantity columns at all (only Particulars/
Posting Date/Debit/Credit/Net Balance) -- a HistoricalDividend needs
`symbol` and cannot be safely constructed from free-text Particulars
without guessing. So this importer never emits ParsedDividend rows; if a
ledger row's Particulars mentions "dividend", a warning is raised instead
of a fabricated dividend record, so a human can check it.
"""

import hashlib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import ClassVar

import openpyxl

from investing_agent.schemas.broker_history import Broker
from investing_agent.services.ingestion.broker_history.base import (
    ParsedCashFlow,
    ParsedStatement,
    ParsedTrade,
    StatementImporter,
    register_importer,
)

TRADEBOOK_HEADER: tuple[str, ...] = (
    "Symbol",
    "ISIN",
    "Trade Date",
    "Exchange",
    "Segment",
    "Series",
    "Trade Type",
    "Auction",
    "Quantity",
    "Price",
    "Trade ID",
    "Order ID",
    "Order Execution Time",
)

LEDGER_HEADER: tuple[str, ...] = (
    "Particulars",
    "Posting Date",
    "Cost Center",
    "Voucher Type",
    "Debit",
    "Credit",
    "Net Balance",
)

# Zerodha ledger Voucher Type -> our CashFlowType. "Book Voucher" is the T+1
# net settlement of trade proceeds (netted across multiple trades, not a
# 1:1 mirror of any single HistoricalTrade row) -- real money movement
# within the account, but not new/withdrawn capital, hence OTHER rather
# than DEPOSIT/WITHDRAWAL.
VOUCHER_TYPE_TO_FLOW_TYPE = {
    "Bank Receipts": "DEPOSIT",
    "Bank Payments": "WITHDRAWAL",
    "Journal Entry": "CHARGE",
    "Book Voucher": "OTHER",
}

_BALANCE_TOLERANCE = Decimal("0.01")


def _trimmed_row(row: tuple) -> tuple:
    """Row values from column B onward (column A is always blank in these
    exports), with trailing Nones dropped."""
    values = list(row[1:])
    while values and values[-1] is None:
        values.pop()
    return tuple(values)


def _find_header_row(ws, signature: tuple[str, ...]) -> int | None:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _trimmed_row(row) == signature:
            return row_idx
    return None


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"Cannot parse decimal value: {value!r}") from exc


@register_importer
class ZerodhaStatementImporter(StatementImporter):
    broker: ClassVar[Broker] = "ZERODHA"

    def parse(self, file_path: Path) -> ParsedStatement:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if (header_row := _find_header_row(ws, TRADEBOOK_HEADER)) is not None:
                return self._parse_tradebook(ws, header_row)
            if (header_row := _find_header_row(ws, LEDGER_HEADER)) is not None:
                return self._parse_ledger(ws, header_row)

        raise ValueError(
            f"Unrecognized Zerodha statement format in {file_path.name!r}: no sheet "
            "matched the known tradebook or ledger column headers. This importer "
            "only handles Console tradebook and funds-ledger exports -- P&L and "
            "holdings statements are not import sources (see module docstring)."
        )

    def _parse_tradebook(self, ws, header_row: int) -> ParsedStatement:
        trades: list[ParsedTrade] = []
        warnings: list[str] = []
        dates: list[date] = []

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = _trimmed_row(row)
            if not values:
                continue
            (
                symbol,
                isin,
                trade_date_raw,
                exchange,
                segment,
                series,
                trade_type,
                auction,
                quantity,
                price,
                trade_id,
                order_id,
                order_execution_time,
            ) = values

            side = trade_type.strip().upper()
            if side not in ("BUY", "SELL"):
                warnings.append(f"trade_id={trade_id}: unrecognized Trade Type {trade_type!r}, skipped")
                continue

            trade_date = date.fromisoformat(str(trade_date_raw))
            dates.append(trade_date)
            trades.append(
                ParsedTrade(
                    dedupe_key=str(trade_id),
                    symbol=str(symbol),
                    isin=str(isin) if isin else None,
                    trade_date=trade_date,
                    trade_datetime=datetime.fromisoformat(str(order_execution_time))
                    if order_execution_time
                    else None,
                    side=side,
                    quantity=_to_decimal(quantity),
                    price=_to_decimal(price),
                    exchange=str(exchange) if exchange else None,
                    segment=str(segment) if segment else None,
                    product=None,
                    order_id=str(order_id) if order_id else None,
                    charges=None,
                    raw={
                        "series": series,
                        "auction": auction,
                        "order_id": order_id,
                        "trade_id": trade_id,
                    },
                )
            )

        return ParsedStatement(
            source_type="zerodha_tradebook_xlsx",
            trades=trades,
            warnings=warnings,
            date_range_start=min(dates) if dates else None,
            date_range_end=max(dates) if dates else None,
        )

    def _parse_ledger(self, ws, header_row: int) -> ParsedStatement:
        cash_flows: list[ParsedCashFlow] = []
        warnings: list[str] = []
        dates: list[date] = []
        running_balance: Decimal | None = None

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = _trimmed_row(row)
            if not values:
                continue
            particulars, posting_date_raw, cost_center, voucher_type, debit, credit, net_balance = values

            net_balance_dec = _to_decimal(net_balance)

            if particulars in ("Opening Balance", "Closing Balance") or not posting_date_raw:
                # Running-balance markers, not real ledger events.
                if particulars == "Opening Balance":
                    running_balance = net_balance_dec
                continue

            debit_dec = _to_decimal(debit)
            credit_dec = _to_decimal(credit)
            amount = credit_dec - debit_dec

            if running_balance is not None:
                running_balance += amount
                if abs(running_balance - net_balance_dec) > _BALANCE_TOLERANCE:
                    warnings.append(
                        f"ledger running-balance mismatch after {posting_date_raw} "
                        f"({particulars!r}): computed {running_balance}, file says {net_balance_dec}"
                    )
                    running_balance = net_balance_dec  # resync so later rows aren't all flagged

            flow_type = VOUCHER_TYPE_TO_FLOW_TYPE.get(str(voucher_type))
            if flow_type is None:
                flow_type = "OTHER"
                warnings.append(f"unrecognized Voucher Type {voucher_type!r} on {posting_date_raw}, mapped to OTHER")

            if "dividend" in str(particulars).lower():
                warnings.append(
                    f"row on {posting_date_raw} mentions 'dividend' in Particulars "
                    f"({particulars!r}) but is not imported as a HistoricalDividend -- "
                    "the ledger has no symbol/quantity columns to build one from safely. "
                    "Verify this row manually."
                )

            posting_date = date.fromisoformat(str(posting_date_raw))
            dates.append(posting_date)

            dedupe_source = f"{posting_date_raw}|{particulars}|{debit_dec}|{credit_dec}"
            cash_flows.append(
                ParsedCashFlow(
                    dedupe_key=hashlib.sha256(dedupe_source.encode()).hexdigest(),
                    flow_date=posting_date,
                    flow_type=flow_type,
                    amount=amount,
                    description=str(particulars),
                    raw={
                        "cost_center": cost_center,
                        "voucher_type": voucher_type,
                        "debit": str(debit_dec),
                        "credit": str(credit_dec),
                        "net_balance": str(net_balance_dec),
                    },
                )
            )

        return ParsedStatement(
            source_type="zerodha_ledger_xlsx",
            cash_flows=cash_flows,
            warnings=warnings,
            date_range_start=min(dates) if dates else None,
            date_range_end=max(dates) if dates else None,
        )
